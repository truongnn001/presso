"""
PressO Python Engine - Excel Handler
=====================================

RESPONSIBILITY:
- Load Excel templates
- Fill cells with provided data
- Export to specified path

ARCHITECTURAL ROLE:
- Stateless document processing
- NO business logic (VAT calculation, contract rules)
- NO database access
- NO external network calls

Reference: PROJECT_DOCUMENTATION.md Section 4.3
"""

import os
import sys
import logging
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, List

# openpyxl for Excel processing
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    Handles Excel document generation.
    
    Stateless handler - each request is independent.
    No business logic - just template filling and export.
    """
    
    # Default template directory (relative to engine)
    TEMPLATE_DIR = Path(__file__).parent.parent / "templates" / "excel"
    
    # Default output directory (system temp)
    OUTPUT_DIR = Path(tempfile.gettempdir()) / "presso" / "excel"
    
    def __init__(self):
        """Initialize Excel handler."""
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not installed - Excel generation unavailable")
        
        # Ensure output directory exists
        self.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        logger.debug(f"ExcelHandler initialized, template_dir={self.TEMPLATE_DIR}")
    
    def handle_export_excel(self, msg_id: Optional[str], params: Dict[str, Any]) -> Dict:
        """
        Handle EXPORT_EXCEL request.
        
        Expected params:
        {
            "template": "basic" (optional, defaults to "basic"),
            "output_filename": "output.xlsx" (optional, auto-generated if not provided),
            "data": {
                "title": "Document Title",
                "date": "2025-12-29",
                "items": [
                    {"name": "Item 1", "quantity": 10, "price": 100.00},
                    {"name": "Item 2", "quantity": 5, "price": 250.00}
                ],
                "notes": "Additional notes"
            }
        }
        
        Returns:
        {
            "id": "...",
            "success": true,
            "result": {
                "file_path": "C:/temp/presso/excel/output_xxx.xlsx",
                "file_size": 12345,
                "template_used": "basic"
            }
        }
        """
        if not OPENPYXL_AVAILABLE:
            return self._make_error(msg_id, "DEPENDENCY_MISSING", 
                "openpyxl library not installed")
        
        try:
            # Extract parameters
            template_name = params.get("template", "basic")
            output_filename = params.get("output_filename")
            data = params.get("data", {})
            
            logger.info(f"Generating Excel: template={template_name}")
            
            # Generate output filename if not provided
            if not output_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                unique_id = str(uuid.uuid4())[:8]
                output_filename = f"presso_{timestamp}_{unique_id}.xlsx"
            
            # Ensure .xlsx extension
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
            
            # Full output path
            output_path = self.OUTPUT_DIR / output_filename
            
            # Generate Excel
            if template_name == "basic":
                self._generate_basic_excel(output_path, data)
            else:
                # Try to load template file
                template_path = self.TEMPLATE_DIR / f"{template_name}.xlsx"
                if template_path.exists():
                    self._generate_from_template(template_path, output_path, data)
                else:
                    return self._make_error(msg_id, "TEMPLATE_NOT_FOUND",
                        f"Template not found: {template_name}")
            
            # Verify output
            if not output_path.exists():
                return self._make_error(msg_id, "GENERATION_FAILED",
                    "Excel file was not created")
            
            file_size = output_path.stat().st_size
            
            logger.info(f"Excel generated: {output_path} ({file_size} bytes)")
            
            return {
                "id": msg_id,
                "success": True,
                "result": {
                    "file_path": str(output_path),
                    "file_name": output_filename,
                    "file_size": file_size,
                    "template_used": template_name,
                    "timestamp": int(datetime.now().timestamp() * 1000)
                }
            }
            
        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            import traceback
            traceback.print_exc(file=sys.stderr)
            return self._make_error(msg_id, "GENERATION_ERROR", str(e))
    
    def _generate_basic_excel(self, output_path: Path, data: Dict[str, Any]):
        """
        Generate a basic Excel document without a template.
        
        This creates a simple document with:
        - Title in A1
        - Date in A2
        - Items table starting at A4
        - Notes at the bottom
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Document"
        
        # Styles
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=11, bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Title
        title = data.get("title", "PressO Document")
        ws.cell(row=row, column=1, value=title).font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        
        # Date
        date_str = data.get("date", datetime.now().strftime("%Y-%m-%d"))
        ws.cell(row=row, column=1, value=f"Date: {date_str}")
        row += 2
        
        # Items table
        items = data.get("items", [])
        
        if items:
            # Headers
            headers = ["#", "Name", "Quantity", "Price", "Total"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Data rows
            grand_total = 0
            for idx, item in enumerate(items, 1):
                name = item.get("name", "")
                quantity = item.get("quantity", 0)
                price = item.get("price", 0)
                total = quantity * price
                grand_total += total
                
                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=2, value=name).border = border
                ws.cell(row=row, column=3, value=quantity).border = border
                cell_price = ws.cell(row=row, column=4, value=price)
                cell_price.border = border
                cell_price.number_format = '#,##0.00'
                cell_total = ws.cell(row=row, column=5, value=total)
                cell_total.border = border
                cell_total.number_format = '#,##0.00'
                row += 1
            
            # Grand total row
            ws.cell(row=row, column=4, value="Grand Total:").font = header_font
            cell_grand = ws.cell(row=row, column=5, value=grand_total)
            cell_grand.font = header_font
            cell_grand.number_format = '#,##0.00'
            row += 2
        
        # Notes
        notes = data.get("notes")
        if notes:
            ws.cell(row=row, column=1, value="Notes:")
            row += 1
            ws.cell(row=row, column=1, value=notes)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Save
        wb.save(output_path)
        logger.debug(f"Basic Excel saved to {output_path}")
    
    def _generate_from_template(self, template_path: Path, output_path: Path, 
                                 data: Dict[str, Any]):
        """
        Generate Excel from a template file.
        
        Template placeholders use format: {{placeholder_name}}
        """
        logger.debug(f"Loading template: {template_path}")
        
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Simple placeholder replacement in all cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value
                    # Replace placeholders
                    for key, val in data.items():
                        if not isinstance(val, (list, dict)):
                            placeholder = "{{" + key + "}}"
                            if placeholder in value:
                                value = value.replace(placeholder, str(val))
                    cell.value = value
        
        # Handle items list if present
        items = data.get("items", [])
        if items:
            # Find the items start row (look for {{items}} or row after headers)
            items_start_row = self._find_items_start_row(ws)
            if items_start_row:
                self._fill_items_table(ws, items_start_row, items)
        
        wb.save(output_path)
        logger.debug(f"Template-based Excel saved to {output_path}")
    
    def _find_items_start_row(self, ws) -> Optional[int]:
        """Find the row where items table should start."""
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "{{items}}" in cell.value or "{{item" in cell.value:
                        return row_idx
        return None
    
    def _fill_items_table(self, ws, start_row: int, items: List[Dict]):
        """Fill items into the table starting at the given row."""
        for idx, item in enumerate(items):
            row = start_row + idx
            col = 1
            
            # Item number
            ws.cell(row=row, column=col, value=idx + 1)
            col += 1
            
            # Item fields
            for key in ["name", "quantity", "price"]:
                if key in item:
                    ws.cell(row=row, column=col, value=item[key])
                col += 1
            
            # Total (quantity * price)
            quantity = item.get("quantity", 0)
            price = item.get("price", 0)
            ws.cell(row=row, column=col, value=quantity * price)
    
    def _make_error(self, msg_id: Optional[str], code: str, message: str) -> Dict:
        """Create an error response."""
        return {
            "id": msg_id,
            "success": False,
            "error": {
                "code": code,
                "message": message
            }
        }

